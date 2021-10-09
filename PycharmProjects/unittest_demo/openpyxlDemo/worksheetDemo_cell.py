import  openpyxl

wb=openpyxl.load_workbook('Book2.xlsx')
##选择要操作的工作表， 返回工作表对象
sheet=wb['test1']
#获取工作表的名称
print(sheet.title)
# 获取工作表中行和列的最值
print(sheet.max_column)
print(sheet.max_row)
print(sheet.min_column)
print(sheet.min_row)
##修改表的名称
sheet.title='test001'
print(sheet.title)
# 返回指定行指定列的单元格信息
print(sheet.cell(row=1,column=2).value)
cell=sheet['B1']
print(cell)
print(cell.row,cell.column,cell.value,cell.coordinate)